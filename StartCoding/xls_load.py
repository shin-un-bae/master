import openpyxl

filename = "empty_book.xlsx"
wb = openpyxl.load_workbook(filename)
sheet = wb.active
for r in range(1, sheet.max_row+1):
    for c in range(1, sheet.max_column+1):
        print(sheet.cell(r, c).value, end=" ")
    #    sheet.cell(r, c).value = f"({r}, {c})"
wb.save(filename)    
