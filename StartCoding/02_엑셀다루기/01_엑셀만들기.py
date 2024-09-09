import openpyxl

wb =openpyxl.Workbook()
ws =wb.create_sheet('오징어게임')    

ws['A1'] ='참가번호'
ws['B1'] ='성명'
ws['A2'] =1
ws['B2'] ='오일남'

# r : 문자형식을 인식시키기 위함
wb.save(r'C:\drmsys\StartCoding\02_엑셀다루기\참가자_data.xlsx')
