import openpyxl
fpath =r'C:\drmsys\StartCoding\02_엑셀다루기\참가자_data.xlsx'
#엑셀 불러오기
wb =openpyxl.load_workbook(fpath)
#엑셀시트 선택
ws =wb['오징어게임']
#데이터 수정하기
ws['a3']=456
ws['b3']='성기훈'

#엑셀 저장하기
wb.save(fpath)
