import  openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx' #생성할 엑셀 파일 이름

ws1 = wb.active
ws1.title = "range names"
# 1번 결과 이미지 참고
for row in range(0, 10):    # row 1 ~ 10 까지 반복하여
    ws1.append(range(10))  # 0~9 까지 append 

# 2번 결과 이미지 참고
ws2 = wb.create_sheet(title="Pi")   # Pi Sheet를 생성 
ws2['F5'] = 3.14                    # F5 셀에 3.14 입력

# 3번 결과 이미지 참고
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

wb.save(filename = dest_filename) # 파일 저장
#[출처] 파이썬으로 엑셀 사용하는 Openpyxl 소개|작성자 은택이